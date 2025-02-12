import os
from django.shortcuts import render
import pandas as pd  # Importa pandas
from django.http import JsonResponse
import requests as rq
from django.conf import settings
from applications.currentstatus.tools import area_ducto_circular, area_inlet_bell, calculate_perdida_choque_codos
from applications.currentstatus.untils import calculo_densidad_aire_sensor, caudal_aire_sensor1, caudal_de_la_frente, presion_dinamica_sensor_1, velocidad_aire_sensor
from applications.getdata.models import Proyecto, SensorsData, VdfData
from django.db.models import Max
from applications.getdata.simulador.simulador import insert_sensor_data
from modules.semaforo import Semaforo


class DataCurrentStatusView:
    
    def __init__(self):
         
         
        #  Inicializa las propiedades como diccionarios vacíos

        self.general = {'FanPerformance': {'status':'' , 'data': []}, 'FanOperation': {'status':'' , 'data': []}}
        self.total_pressure = {'FanPerformance': {'status':'' , 'data': []}, 'FanOperation': {'status':'' , 'data': []}}
        self.static_pressure = {'FanPerformance': {'status':'' , 'data': []}, 'FanOperation': {'status':'' , 'data': []}}
        self.power = {'FanPerformance': {'status':'' , 'data': []}, 'FanOperation': {'status':'' , 'data': []}}
    
    def add_measurement(self, property_name, fan_type, status, measurement):

        # Añade una medida a la propiedad correspondiente
        if property_name in ['general', 'total_pressure', 'static_pressure', 'power']:
            if fan_type in ['FanPerformance', 'FanOperation']:
                getattr(self, property_name)[fan_type]['data']= measurement
                getattr(self, property_name)[fan_type]['status'] = status
            else:
                print("Error: Fan type must be 'FanPerformance' or 'FanOperation'")
        else:
            print("Error: Invalid property name")
    
    def to_dict(self):
        # Retorna un diccionario con todas las propiedades
        return {
            'general': self.general,
            'total_pressure': self.total_pressure,
            'static_pressure': self.static_pressure,
            'power': self.power
        }
   

def currentstatus(request):
    
    if request.method == 'GET':
        csv_file_path = os.path.join(settings.MEDIA_ROOT, 'datos.csv')
        df = pd.read_csv(csv_file_path)
        context = {}
        context['project'] = Proyecto.objects.all().last()
        # Ejemplo de uso
        data_view = DataCurrentStatusView()

        # Añadiendo medidas
        data_view.add_measurement('general', 'FanPerformance','green',df[["q1", "pt1"]].to_dict(orient='records') )
        data_view.add_measurement('general', 'FanOperation','yellow', df[["q1", "pt1"]].to_dict(orient='records') )

        data_view.add_measurement('total_pressure', 'FanPerformance','red', df[["q1", "pt1"]].to_dict(orient='records'))
        data_view.add_measurement('total_pressure', 'FanOperation', 'yellow',  df[["q1", "pt1"]].to_dict(orient='records') )

        data_view.add_measurement('static_pressure', 'FanPerformance','red',  df[["q1", "pt1"]].to_dict(orient='records'))
        data_view.add_measurement('static_pressure', 'FanOperation','green',  df[["q1", "pt1"]].to_dict(orient='records'))

        data_view.add_measurement('power', 'FanPerformance','red', df[["q1", "pt1"]].to_dict(orient='records'))
        data_view.add_measurement('power', 'FanOperation', 'yellow', df[["q1", "pt1"]].to_dict(orient='records'))

        context['data']=data_view.to_dict()
        return render(request, 'currentStatus.html', context)
    
    # Si la solicitud no es un POST, simplemente renderiza la página sin datos
    return render(request, 'currentStatus.html')


def get_recent_data(request):
    '''

    '''
    if request.method == 'GET':
        # VARIABLES DEL SENSOR 
        latest_record_sensors = SensorsData.objects.using('sensorDB').aggregate(Max('id'))
        max_id_sensors = latest_record_sensors['id__max']
        item_sensors = SensorsData.objects.using('sensorDB').get(id=max_id_sensors)
        variables = {}
        insert_sensor_data()
        latest_record_vdf = VdfData.objects.using('sensorDB').aggregate(Max('id'))
        max_id_vdf = latest_record_vdf['id__max']

        # calcular la perdida por choque
        project = Proyecto.objects.all().order_by('id').last() 
        caracteristicas = project.ventilador.accesorios.all()


        variables['vmm'] = project.ventilador.vmm
        variables['amm'] = project.ventilador.amm
        variables['nmm'] = project.ventilador.nmm
        
        #  CALCULANDO LA DENSIDAD DEL AIRE 
        tbs = item_sensors.lc # Tbs1
        hr =  item_sensors.qf # HRs1
        
        q2 = item_sensors.q2 # Tbs2
        pt1 = item_sensors.pt1 # HRs2
        
        q1 = item_sensors.q1  # Pbs1

        semaforo = Semaforo(request)
        semaforo.encender(project)
        tbs1 = round(semaforo.calculate_tbh(tbs, hr),1)
        tbh1 = round(semaforo.calculate_tbh(q2, pt1),1)
        
        calculador_densidad_aire_s1 = calculo_densidad_aire_sensor(tbs1, tbh1, q1)
        
        item_vdf = VdfData.objects.using('sensorDB').get(id=max_id_vdf)
        variables['densidad'] = item_sensors.densidad1
        # calcular la densidad:
        densidad = calculador_densidad_aire_s1.densidad_del_aire()
        
        mid_densidad = densidad/2
        # TODO estas valores son solo una prueba de las funciones
        caudal_del_ventilador = semaforo.calculate_Q1()
        Qf = caudal_de_la_frente(semaforo.calculate_Q2(), semaforo.leakage_coefficient_v4(), item_sensors.pt2, project.ducto.Ldsf )
        
        # cambio de valores 
        variables ['caudal_del_ventilador'] = caudal_del_ventilador
        
        Q_codo_1 = caudal_del_ventilador**2
        area_inlet_bell_val = area_inlet_bell(project)
        Area_ventilador = 3.14159 * (project.ventilador.amm/2000)**2
        variables['area_inlet_bell_val']  = area_inlet_bell_val
        variables['Area_ventilador']  = Area_ventilador
        
        presion_dinamica_entrada_Pa = mid_densidad * Q_codo_1 / (area_inlet_bell_val*area_inlet_bell_val)
  
        variables['presion_dinamica_entrada_Pa'] = presion_dinamica_entrada_Pa

        Presion_dinamica_ventilador_Pa = mid_densidad * Q_codo_1 / (Area_ventilador*Area_ventilador)
        variables['Presion_dinamica_ventilador_Pa'] = Presion_dinamica_ventilador_Pa
        '''
           La perdida de choque de accesorios es la presion de choque de accesorios por la presion dinamica
        '''
        sumatoria_choque_accesorios = 0
        for caracteristica in caracteristicas:
            sumatoria_choque_accesorios += caracteristica.factor_choque*presion_dinamica_entrada_Pa

        # calcular perdida por choque de los codos
        total_codos = project.codos
        variables['sumatoria_choque_accesorios'] = sumatoria_choque_accesorios

        try:
            vars, perdidas_choque_codos = calculate_perdida_choque_codos(
                        total_codos=total_codos, 
                        mid_densidad=mid_densidad, 
                        Q1=caudal_del_ventilador,
                        project=project, 
                        Qf = Qf
                        )
            variables["perdida_choque_codos"] = perdidas_choque_codos

        except TypeError as e:
            context = {}
            context["status"] = "error"
            context["message"] = f"No se pudo realizar el calculo de calcular choque de codos: {e}, verifique el valor de diametro del ducto"
            return JsonResponse(context, safe=False)

        try:
            area_ducto_circular_ = area_ducto_circular(project)
        except TypeError as e:
            context = {}
            context["status"] = "error"
            context["message"] = f"No se pudo realizar el calculo: {e}, verifique el valor de diametro del ducto"

            return JsonResponse(context, safe=False)

        perdida_choque_salida_ducto_circular = mid_densidad*(Qf*Qf/(area_ducto_circular_*area_ducto_circular_))
        perdida_choque_salida_ducto_ovalado = mid_densidad*Qf*Qf/(project.ducto.area*project.ducto.area)

        perdida_choque_salida_ducto = None 
        if project.ducto.t_ducto == "ovalado":
            perdida_choque_salida_ducto = perdida_choque_salida_ducto_ovalado
        elif project.ducto.t_ducto == "circular":
            perdida_choque_salida_ducto = perdida_choque_salida_ducto_circular
        
        if perdida_choque_salida_ducto is None:
            raise ValueError("No se pudo identificar el tipo de ducto")

        variables['perdida_choque_salida_ducto'] = perdida_choque_salida_ducto
        variables['perdida_choque_salida_ducto2'] = perdida_choque_salida_ducto
        #a -> resistencia
        presion_total = item_sensors.ps1 
        presion_estatica_ventilador = round(presion_total - presion_dinamica_entrada_Pa, 0)
        # este calculo obtiene el valor  adecuado independientemente del tipo de ducto, es decir funciona para circular y ovalado
        perdida_choque_total_sistema_ducto = perdidas_choque_codos +sumatoria_choque_accesorios+perdida_choque_salida_ducto

        variables['perdida_choque_total_sistema_ducto'] = perdida_choque_total_sistema_ducto
        #perdida_choque_total_sistema_ducto_ovalado_Pa = perdidas_choque_codos +sumatoria_choque_accesorios + perdida_choque_salida_ducto_ovalado
        perdidas_friccionales = round(presion_estatica_ventilador - perdida_choque_total_sistema_ducto, 2)
        variables['perdidas_friccionales'] = perdidas_friccionales

        presion_dinamica = round(presion_dinamica_sensor_1(item_sensors.ps1, item_sensors.densidad1 ),0)
        # calculando el caudal del aire sensor 1
        velocidad_aire_sensor1 = velocidad_aire_sensor(presion_dinamica, calculador_densidad_aire_s1.densidad_del_aire())
        
        area_ducto = project.ducto.area
        data = {
            "ps1": round(item_sensors.ps1, 2),
            "qf": round(item_sensors.qf, 2),
            "q1": caudal_aire_sensor1(velocidad_aire_sensor1, area_ducto),
            "pt1": round(item_sensors.pt1, 2),
            "densidad1": round(item_sensors.densidad1, 2),
            "powerc": round(item_vdf.powerc, 2),
            "fref": round(item_vdf.fref, 2),
            "frequency_ratio_1": round((item_vdf.freal / item_vdf.fref) * 100, 2),
            "frequency_ratio_2": round((item_vdf.freal / item_vdf.fref) * 100, 2),
            "powerc_duplicate": round(item_vdf.powerc, 2),
        }
        
        context = {}
        context["data"] = data
        presion_dinamica = item_sensors.ps1 - item_sensors.densidad1
        context["presion_estatica"] = round(presion_estatica_ventilador, 1)
        context["presion_dinamica"] = round(presion_dinamica, 1)
        context["perdida_de_choque"] = round(perdida_choque_total_sistema_ducto, 1)
        context["perdidas_friccionales"] = round(perdidas_friccionales, 1)
        context['variables'] = variables
        return JsonResponse(context, safe=False)
    
def update_frequency(request):
   if request.method == 'GET':
      newFref = request.GET.get('frecuency')
      url = f"http://localhost:1880/update-frequency?frecuency={newFref}"
      try:
            #Enviar al endpoint del Node-Red
            response = rq.get(url)
      except rq.exceptions.RequestException as e:
            print(f"Error updating frequency: {e}")

   return JsonResponse('Frecuencia Ref Actualizada', safe=False)


def Excel(request):
    import datetime
    from django.utils import timezone
    import openpyxl
    from django.http import HttpResponse

    now = timezone.now() - datetime.timedelta(days=1) #Obtencion de Timezone menos 24horas

    timestamp = datetime.datetime.now()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Datos_{timestamp}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'VDF'

    # Write header row
    header = ['Frecuencia referencia', 'Frecuencia Real', 'VF', 'IF', 'power','powerc','rpm']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = VdfData.objects.filter(ts__gte=now).values_list('fref', 'freal', 'vf','oc','power','powerc','rpm')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    ############
            
    workbook.create_sheet('Sensores')
    workbook.active = workbook['Sensores']
    worksheet = workbook.active
    
    # Write header row
    header = ['pt2', 'ps2', 'densidad2','q2','pt1','ps1','densidad1','q1','lc','qf','k','tbs','hr','tbh','tgbh']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = SensorsData.objects.filter(ts__gte=now).values_list('pt2', 'ps2', 'densidad2','q2','pt1','ps1','densidad1','q1','lc','qf','k','tbs','hr','tbh','tgbh')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    

    workbook.save(response)

    return response